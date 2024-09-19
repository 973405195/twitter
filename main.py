import requests,os
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from requests.cookies import RequestsCookieJar
from openpyxl import load_workbook
from openpyxl import Workbook
from time import sleep
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import string
import zipfile
from concurrent.futures import ThreadPoolExecutor
import shutil

# 打包Google代理插件
def create_proxyauth_extension(proxy_host, proxy_port, proxy_username, proxy_password, scheme='http', plugin_path=None):
    if plugin_path is None:
        # 插件地址
        plugin_path = 'Selenium-Chrome-HTTP-Private-Proxy-master.zip'

    manifest_json = """
        {
            "version": "1.0.0",
            "manifest_version": 2,
            "name": "Chrome Proxy",
            "permissions": [
                "proxy",
                "tabs",
                "unlimitedStorage",
                "storage",
                "<all_urls>",
                "webRequest",
                "webRequestBlocking"
            ],
            "background": {
                "scripts": ["background.js"]
            },
            "minimum_chrome_version":"22.0.0"
        }
        """

    background_js = string.Template(
        """
        var config = {
                mode: "fixed_servers",
                rules: {
                  singleProxy: {
                    scheme: "${scheme}",
                    host: "${host}",
                    port: parseInt(${port})
                  },
                  bypassList: ["foobar.com"]
                }
              };

        chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});

        function callbackFn(details) {
            return {
                authCredentials: {
                    username: "${username}",
                    password: "${password}"
                }
            };
        }

        chrome.webRequest.onAuthRequired.addListener(
                    callbackFn,
                    {urls: ["<all_urls>"]},
                    ['blocking']
        );
        """
    ).substitute(
        host=proxy_host,
        port=proxy_port,
        username=proxy_username,
        password=proxy_password,
        scheme=scheme,
    )
    with zipfile.ZipFile(plugin_path, 'w') as zp:
        zp.writestr("manifest.json", manifest_json)
        zp.writestr("background.js", background_js)

    return plugin_path

# 填写主机地址，端口，账号，密码
proxyauth_plugin_path = create_proxyauth_extension(
    proxy_host="na.fq2alntp.lunaproxy.net",
    proxy_port=12233,
    proxy_username="user-nishui6-region-br",
    proxy_password="123456"
)



dailimoshi = input('是否需要代理，输入0需要，输入1需要代理。输入后回车：')
xiancheng_num = input('请输入本次需要开启的线程：')

cookie_strs = []
names = []
passwords = []
emails = []
email_passwords = []
news_data= []
errors_data = []
driver_nums = []

def main(cookie_str,zh_name,password,email,email_password,driver_num):
    try:
        print(f'第{driver_num}个账号:{zh_name} 开始获取cookies')
        options = webdriver.ChromeOptions()
        options.add_experimental_option('useAutomationExtension', False)
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_argument('--start-maximized')
        options.add_argument('disable-infobars')
        prefs = {'profile.default_content_setting_values': {'images': 2, 'permissions.default.stylesheet': 2}}
        options.add_experimental_option('prefs', prefs)
        options.add_argument("--start-maximized")
        if dailimoshi == '1':
            options.add_extension('Selenium-Chrome-HTTP-Private-Proxy-master.zip')
        else:
            pass
            # options.add_argument('--headless')  # 无头
        # options.add_argument(
        #     'User-Agent="Mozilla/5.0 (Linux; Android 8.0.0; LG-H870S) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.101 Mobile Safari/537.36;"')

        driver = webdriver.Chrome(options=options)
        driver.get('https://x.com/')
        cookies = {}
        for line in cookie_str.split(";"):
            # print(line)
            if line.find("=") != -1:
                name, value = line.strip().split("=", 1)
                cookies[name] = value

        # print(cookies)
        for cookie in cookies:
            # print(cookie)
            # print(cookies[cookie])
            driver.add_cookie({
                "domain": ".x.com",
                "name": cookie,
                "value": cookies[cookie],
                "path": '/',
                "expires": None
            })
        driver.refresh()
        sleep(2)
        driver.execute_script("""
                                        let videos = document.querySelectorAll('video');
                                        videos.forEach(video => {
                                            video.autoplay = false;
                                            video.load();
                                        });
                                    """)
        driver_cookies = driver.get_cookies()

        driver_cookie_str = ''
        for cookie in driver_cookies:
            driver_cookie_str += f"{cookie['name']}={cookie['value']}; "

        # 可以使用append方法插入一行数据
        news_data.append([zh_name, password, email, email_password, driver_cookie_str])

        print('获取成功，执行下一个')

        driver.quit()
        driver_num+=1
    except Exception as e:
        print(e)
        print(f'{zh_name}的cookies获取失败，执行下一个')
        errors_data.append([zh_name, password, email, email_password])
        driver_num += 1




try:
    # 加载工作簿
    wb = load_workbook('ck.xlsx')
    sheet = wb.active
    # 读取所有非None值的行
    non_none_rows = []
    for row in sheet.iter_rows():
        non_none_row = [cell.value for cell in row if cell.value is not None]
        if non_none_row:
            non_none_rows.append(non_none_row)

    #成功存储
    cr_wb = Workbook()
    # 选择默认的工作表
    cr_ws = cr_wb.active

    #失败存储
    cw_wb = Workbook()
    # 选择默认的工作表
    cw_ws = cw_wb.active

    # 打印结果
    num = 1
    for row in non_none_rows:
        try:
            print(f'开始遍历第{num}行账号')
            # 登录URL
            login_url = 'http://x.com/home'
            base_url = 'http://x.com'
            # 登录所需的Token信息
            data = {
                'auth_token': row[4]
            }
            # 发送POST请求进行登录
            headers = {
                'User-Agent':'Mozilla/5.0 (Linux; Android 8.0.0; LG-H870S) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.101 Mobile Safari/537.36;'
            }
            proxiess = {
                'http': 'http://127.0.0.1:9098',
                'https': 'https://127.0.0.1:9098'
            }

            response = requests.post(base_url, data=data, stream=True,headers=headers)
            # 检查是否登录成功
            if response.ok:
                # 获取登录后的cookies
                cookie_str = response.headers.get("set-cookie")
                # #获取登录后的token
                au_token = response.json().get("token")
                test_cookies = response.cookies

                cookie_str += f"'; auth_token={data['auth_token']}"
                cookie_strs.append(cookie_str)
                print(cookie_str)
                print(au_token)
                names.append(row[0])
                passwords.append(row[1])
                emails.append(row[2])
                email_passwords.append(row[3])
                num+=1

            else:
                cw_ws.append([row[0], row[1], row[2], row[3]])
                cw_wb.save("error.xlsx")
                print(f'第{num}行账号遍历失败')
                num+=1

        except Exception as e :
            print(e)
            print(f'第{num}行账号遍历报错')
            num+=1
            cw_ws.append([row[0], row[1], row[2], row[3]])
            cw_wb.save("error.xlsx")



    # for i in range(0,len(names)):
    #     print(i+1)
    #     driver_nums.append(i+1)
    # with ThreadPoolExecutor(max_workers=int(xiancheng_num)) as executor:
    #     executor.map(main, cookie_strs,names,passwords,emails,email_passwords,driver_nums)
    # os.system('taskkill /im chromedriver.exe /F')
    # os.system('taskkill /im chrome.exe /F')
    #
    # print('开始执行xlsx读写程序')
    # for n in news_data:
    #     cr_ws.append(n)
    # for e in errors_data:
    #     cw_ws.append(e)
    # # 保存工作簿到文件
    # cr_wb.save("new.xlsx")
    # cw_wb.save('error.xlsx')
    # print('开始清除系统临时缓存,耐心等待')
    # shutil.rmtree(r'C:\Windows\SystemTemp', ignore_errors=True)
    # try:
    #     os.mkdir(r'C:\Windows\SystemTemp')
    # except FileExistsError:
    #     print('文件夹已存在')
    # print('程序执行完毕')

except Exception as e:
    print(e)



